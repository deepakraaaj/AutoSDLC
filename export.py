from io import BytesIO
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from schemas import GenerationOutput


def get_priority_color(priority: str) -> str:
    """Return hex color for priority level."""
    colors = {
        "critical": "8B0000",  # dark red
        "high": "FF8C00",      # dark orange
        "medium": "DAA520",    # goldenrod
        "low": "228B22"        # forest green
    }
    return colors.get(priority, "808080")  # default gray


def generate_excel(output: GenerationOutput) -> bytes:
    """Generate Excel .xlsx file from GenerationOutput. Returns bytes."""
    wb = Workbook()
    wb.remove(wb.active)

    # Define styles
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="1F6FEB", end_color="1F6FEB", fill_type="solid")
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    wrap = Alignment(wrap_text=True, vertical='top')

    # Sheet 0: Epics
    ws_epics = wb.create_sheet("Epics", 0)
    headers_epics = ["ID", "Title", "Description", "Feature Area", "Priority", "Status"]
    for col, header in enumerate(headers_epics, 1):
        cell = ws_epics.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = border
        cell.alignment = Alignment(horizontal='center', vertical='center')

    for row_idx, epic in enumerate(output.epics, 2):
        ws_epics.cell(row=row_idx, column=1, value=epic.id).border = border
        ws_epics.cell(row=row_idx, column=2, value=epic.title).border = border
        ws_epics.cell(row=row_idx, column=3, value=epic.description).border = border
        ws_epics.cell(row=row_idx, column=3).alignment = wrap
        ws_epics.cell(row=row_idx, column=4, value=epic.feature_area).border = border

        priority_cell = ws_epics.cell(row=row_idx, column=5, value=epic.priority)
        priority_cell.border = border
        priority_color = get_priority_color(epic.priority)
        priority_cell.fill = PatternFill(start_color=priority_color, end_color=priority_color, fill_type="solid")
        priority_cell.font = Font(color="FFFFFF", bold=True)

        ws_epics.cell(row=row_idx, column=6, value=epic.status).border = border

    ws_epics.column_dimensions['A'].width = 8
    ws_epics.column_dimensions['B'].width = 20
    ws_epics.column_dimensions['C'].width = 25
    ws_epics.column_dimensions['D'].width = 15
    ws_epics.column_dimensions['E'].width = 12
    ws_epics.column_dimensions['F'].width = 12
    ws_epics.row_dimensions[1].height = 25

    # Sheet 1: User Stories
    ws_stories = wb.create_sheet("User Stories", 1)
    headers_stories = ["ID", "Epic ID", "Title", "As A", "I Want", "So That", "Acceptance Criteria", "Feature Area", "Size", "Priority", "Status", "Confidence"]
    for col, header in enumerate(headers_stories, 1):
        cell = ws_stories.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = border
        cell.alignment = Alignment(horizontal='center', vertical='center')

    for row_idx, story in enumerate(output.stories, 2):
        ws_stories.cell(row=row_idx, column=1, value=story.id).border = border
        ws_stories.cell(row=row_idx, column=2, value=story.epic_id or "—").border = border
        ws_stories.cell(row=row_idx, column=3, value=story.title).border = border
        ws_stories.cell(row=row_idx, column=4, value=story.as_a).border = border
        ws_stories.cell(row=row_idx, column=4).alignment = wrap
        ws_stories.cell(row=row_idx, column=5, value=story.i_want).border = border
        ws_stories.cell(row=row_idx, column=5).alignment = wrap
        ws_stories.cell(row=row_idx, column=6, value=story.so_that).border = border
        ws_stories.cell(row=row_idx, column=6).alignment = wrap
        ac_text = "\n".join(f"• {ac}" for ac in story.acceptance_criteria)
        ws_stories.cell(row=row_idx, column=7, value=ac_text).border = border
        ws_stories.cell(row=row_idx, column=7).alignment = wrap
        ws_stories.cell(row=row_idx, column=8, value=story.feature_area).border = border
        ws_stories.cell(row=row_idx, column=9, value=story.size).border = border

        priority_cell = ws_stories.cell(row=row_idx, column=10, value=story.priority)
        priority_cell.border = border
        priority_color = get_priority_color(story.priority)
        priority_cell.fill = PatternFill(start_color=priority_color, end_color=priority_color, fill_type="solid")
        priority_cell.font = Font(color="FFFFFF", bold=True)

        ws_stories.cell(row=row_idx, column=11, value=story.status).border = border
        ws_stories.cell(row=row_idx, column=12, value=story.confidence).border = border

    ws_stories.column_dimensions['A'].width = 6
    ws_stories.column_dimensions['B'].width = 8
    ws_stories.column_dimensions['C'].width = 18
    ws_stories.column_dimensions['D'].width = 15
    ws_stories.column_dimensions['E'].width = 18
    ws_stories.column_dimensions['F'].width = 18
    ws_stories.column_dimensions['G'].width = 25
    ws_stories.column_dimensions['H'].width = 12
    ws_stories.column_dimensions['I'].width = 8
    ws_stories.column_dimensions['J'].width = 12
    ws_stories.column_dimensions['K'].width = 12
    ws_stories.column_dimensions['L'].width = 12
    ws_stories.row_dimensions[1].height = 25

    # Sheet 2: Developer Tasks
    ws_tasks = wb.create_sheet("Developer Tasks", 2)
    headers_tasks = ["ID", "Story ID", "Title", "Description", "Definition of Done", "Estimate (hrs)", "Dependencies", "Priority", "Status", "Assignee", "Confidence"]
    for col, header in enumerate(headers_tasks, 1):
        cell = ws_tasks.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = border
        cell.alignment = Alignment(horizontal='center', vertical='center')

    for row_idx, task in enumerate(output.tasks, 2):
        ws_tasks.cell(row=row_idx, column=1, value=task.id).border = border
        ws_tasks.cell(row=row_idx, column=2, value=task.story_id or "—").border = border
        ws_tasks.cell(row=row_idx, column=3, value=task.title).border = border
        ws_tasks.cell(row=row_idx, column=4, value=task.description).border = border
        ws_tasks.cell(row=row_idx, column=4).alignment = wrap
        ws_tasks.cell(row=row_idx, column=5, value=task.definition_of_done).border = border
        ws_tasks.cell(row=row_idx, column=5).alignment = wrap
        ws_tasks.cell(row=row_idx, column=6, value=task.estimate_hours).border = border
        deps_text = "\n".join(f"• {d}" for d in task.dependencies) if task.dependencies else ""
        ws_tasks.cell(row=row_idx, column=7, value=deps_text).border = border
        ws_tasks.cell(row=row_idx, column=7).alignment = wrap

        priority_cell = ws_tasks.cell(row=row_idx, column=8, value=task.priority)
        priority_cell.border = border
        priority_color = get_priority_color(task.priority)
        priority_cell.fill = PatternFill(start_color=priority_color, end_color=priority_color, fill_type="solid")
        priority_cell.font = Font(color="FFFFFF", bold=True)

        ws_tasks.cell(row=row_idx, column=9, value=task.status).border = border
        ws_tasks.cell(row=row_idx, column=10, value=task.assignee or "Unassigned").border = border
        ws_tasks.cell(row=row_idx, column=11, value=task.confidence).border = border

    ws_tasks.column_dimensions['A'].width = 6
    ws_tasks.column_dimensions['B'].width = 8
    ws_tasks.column_dimensions['C'].width = 18
    ws_tasks.column_dimensions['D'].width = 22
    ws_tasks.column_dimensions['E'].width = 22
    ws_tasks.column_dimensions['F'].width = 12
    ws_tasks.column_dimensions['G'].width = 22
    ws_tasks.column_dimensions['H'].width = 12
    ws_tasks.column_dimensions['I'].width = 12
    ws_tasks.column_dimensions['J'].width = 15
    ws_tasks.column_dimensions['K'].width = 12
    ws_tasks.row_dimensions[1].height = 25

    # Write to BytesIO
    output_file = BytesIO()
    wb.save(output_file)
    output_file.seek(0)
    return output_file.getvalue()
